from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelReport:
    def create_summary(self, records, output_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        headers = [
            "Generated Time",
            "Hostname",
            "Model",
            "Serial",
            "DDOS",
            "Pre-Comp Used (TiB)",
            "Post-Comp Used (TiB)",
            "Post-Comp Size (TiB)",
            "Available (TiB)",
            "Usage %",
            "Growth (TiB)",
            "Growth %",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        records = sorted(
            records,
            key=lambda x: x.get("generated_time", "")
        )

        previous_used = None
        row = 2

        for r in records:
            current_used = r.get("post_comp_used_tib")

            growth_tib = None
            growth_pct = None

            if previous_used is not None and current_used is not None:
                growth_tib = round(current_used - previous_used, 2)

                if previous_used != 0:
                    growth_pct = round((growth_tib / previous_used) * 100, 2)

            previous_used = current_used

            ws.cell(row=row, column=1).value = r.get("generated_time")
            ws.cell(row=row, column=2).value = r.get("hostname")
            ws.cell(row=row, column=3).value = r.get("model")
            ws.cell(row=row, column=4).value = r.get("serial")
            ws.cell(row=row, column=5).value = r.get("ddos")
            ws.cell(row=row, column=6).value = r.get("pre_comp_used_tib")
            ws.cell(row=row, column=7).value = r.get("post_comp_used_tib")
            ws.cell(row=row, column=8).value = r.get("post_comp_size_tib")
            ws.cell(row=row, column=9).value = r.get("post_comp_avail_tib")
            ws.cell(row=row, column=10).value = r.get("use_pct")
            ws.cell(row=row, column=11).value = growth_tib
            ws.cell(row=row, column=12).value = growth_pct

            row += 1

        for col in range(1, 13):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

        ws.freeze_panes = "A2"

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)